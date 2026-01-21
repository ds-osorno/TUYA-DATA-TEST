#!/usr/bin/env python3
"""Pipeline para calcular rachas desde el Excel hacia SQLite.

Este script carga las hojas `historia` y `retiros`, aplica controles básicos
y ejecuta la consulta SQL que devuelve la racha por cliente.
"""

from __future__ import annotations

import argparse
import csv
import logging
import sqlite3
from dataclasses import dataclass
from datetime import datetime, date
from pathlib import Path
from typing import Any, List, Optional, Tuple

from openpyxl import load_workbook


LOG = logging.getLogger("rachas")


HISTORIA_COLS = ["identificacion", "corte_mes", "saldo"]
RETIROS_COLS = ["identificacion", "fecha_retiro"]


@dataclass(frozen=True)
class HistoriaRow:
    identificacion: str
    corte_mes: str
    saldo: int


@dataclass(frozen=True)
class RetiroRow:
    identificacion: str
    fecha_retiro: Optional[str]


def setup_logging(verbose: bool) -> None:
    lvl = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(level=lvl, format="%(levelname)s: %(message)s")


def parse_args() -> argparse.Namespace:
    # DB por defecto en la carpeta del script
    default_db = str(Path(__file__).parent / "rachas.db")

    p = argparse.ArgumentParser(
        description="Rachas -> SQLite",
        epilog="Ejemplo: python rachas_pipeline.py --fecha-base 2024-12-15 --n 3",
    )
    p.add_argument("--excel", default="Prueba Tecnica.xlsx", help="Ruta al Excel")
    p.add_argument("--db", default=default_db, help="Archivo sqlite (default: rachas.db en carpeta del script)")
    p.add_argument("--fecha-base", required=True, dest="fecha_base", help="YYYY-MM-DD")
    p.add_argument("--n", required=True, type=int, help="Número mínimo de meses consecutivos para considerar una racha válida")
    p.add_argument("--out", default="3_rachas_sql/resultados_rachas.csv", help="CSV de salida (default: resultados_rachas.csv)")
    p.add_argument("--verbose", action="store_true")
    return p.parse_args()


def read_sql(path: Path) -> str:
    return path.read_text(encoding="utf-8").strip()


def parse_date(v: Any) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.date().isoformat() if isinstance(v, datetime) else v.isoformat()
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s).date().isoformat()
    except ValueError:
        return None


def normalize_month_end(d: str) -> str:
    dt = datetime.strptime(d, "%Y-%m-%d").date()
    nxt = date(dt.year + 1, 1, 1) if dt.month == 12 else date(dt.year, dt.month + 1, 1)
    return (date.fromordinal(nxt.toordinal() - 1)).isoformat()


def to_int(v: Any) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(round(v))
    s = str(v).strip().replace(",", "").replace(".", "")
    return int(s) if s.isdigit() else None


def check_header(row: Tuple[Any, ...], expected: List[str], sheet: str) -> None:
    got = [str(c).strip().lower() if c is not None else "" for c in row]
    exp = [e.lower() for e in expected]
    if got[: len(exp)] != exp:
        raise ValueError(f"Hoja {sheet}: encabezados inválidos: {row[:len(exp)]}")


def read_excel(path: Path) -> Tuple[List[HistoriaRow], List[RetiroRow]]:
    wb = load_workbook(path, data_only=True)
    if "historia" not in wb.sheetnames or "retiros" not in wb.sheetnames:
        raise ValueError("Faltan hojas 'historia' o 'retiros'")

    # historia
    ws = wb["historia"]
    it = ws.iter_rows(values_only=True)
    header = next(it)
    check_header(header, HISTORIA_COLS, "historia")

    historia_rows = []
    skipped = 0
    negatives = 0
    for i, r in enumerate(it, start=2):
        id_ = str(r[0] or "").strip()
        corte = parse_date(r[1])
        saldo = to_int(r[2])
        if not id_ or not corte or saldo is None:
            skipped += 1
            LOG.warning(f"historia fila {i}: omitida")
            continue
        if saldo < 0:
            negatives += 1
            saldo = 0
        historia_rows.append(HistoriaRow(id_, normalize_month_end(corte), saldo))

    # retiros
    ws = wb["retiros"]
    it = ws.iter_rows(values_only=True)
    header = next(it)
    check_header(header, RETIROS_COLS, "retiros")

    retiros = {}
    skipped_retiros = 0
    for i, r in enumerate(it, start=2):
        id_ = str(r[0] or "").strip()
        fecha = parse_date(r[1])
        if not id_:
            skipped_retiros += 1
            LOG.warning(f"retiros fila {i}: identificacion vacía")
            continue
        retiros[id_] = RetiroRow(id_, fecha)

    LOG.info(f"historia: {len(historia_rows)} registros cargados (omitidos: {skipped}, negativos: {negatives})")
    LOG.info(f"retiros: {len(retiros)} registros cargados (omitidos: {skipped_retiros})")

    return historia_rows, list(retiros.values())


def init_db(conn: sqlite3.Connection, ddl: str) -> None:
    conn.executescript(ddl)
    conn.commit()


def load_data(conn: sqlite3.Connection, historia: List[HistoriaRow], retiros: List[RetiroRow]) -> None:
    conn.execute("BEGIN;")
    conn.executemany("INSERT OR REPLACE INTO historia_saldos(identificacion,corte_mes,saldo) VALUES (?,?,?)",
                     [(h.identificacion, h.corte_mes, h.saldo) for h in historia])
    conn.executemany("INSERT OR REPLACE INTO retiros(identificacion,fecha_retiro) VALUES (?,?)",
                     [(r.identificacion, r.fecha_retiro) for r in retiros])
    conn.commit()


def run_query(conn: sqlite3.Connection, query: str, fecha_base: str, n: int):
    cur = conn.cursor()
    cur.execute(query, {"fecha_base": fecha_base, "n": n})
    return cur.fetchall()


def write_csv(path: Path, rows: List[Tuple[Any, ...]]):
    hdr = ["identificacion", "racha", "fecha_fin", "nivel"]
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(hdr)
        w.writerows(rows)


def main():
    args = parse_args()
    setup_logging(args.verbose)

    excel = Path(args.excel)
    if not excel.exists():
        raise FileNotFoundError(f"No se encontró: {excel}")

    # fecha_base valida
    try:
        datetime.strptime(args.fecha_base, "%Y-%m-%d")
    except ValueError:
        raise ValueError("fecha_base debe estar en formato YYYY-MM-DD")

    LOG.info(f"Procesando Excel: {excel}")
    LOG.info(f"Parámetros: fecha_base={args.fecha_base}, n={args.n}")

    sql_dir = Path(__file__).parent / "sql"
    ddl = read_sql(sql_dir / "ddl.sql")
    query = read_sql(sql_dir / "query_rachas.sql")

    historia, retiros = read_excel(excel)

    conn = sqlite3.connect(args.db)
    try:
        init_db(conn, ddl)
        load_data(conn, historia, retiros)
        rows = run_query(conn, query, args.fecha_base, args.n)
        LOG.info(f"✓ {len(rows)} clientes con rachas >= {args.n} meses")

        print("\n" + "="*80)
        print(f"RESULTADOS: Top 20 rachas (de {len(rows)} totales)")
        print("="*80)
        for r in rows[:20]:
            print(f"  {r[0]:<20} | racha: {r[1]:>2} meses | fin: {r[2]} | nivel: {r[3]}")
        if len(rows) > 20:
            print(f"  ... y {len(rows)-20} clientes más")
        print("="*80 + "\n")

        write_csv(Path(args.out), rows)
        LOG.info(f"✓ Resultados guardados en: {args.out}")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
